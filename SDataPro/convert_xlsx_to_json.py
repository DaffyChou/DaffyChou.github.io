#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
convert_xlsx_to_json.py
========================
將船訊網 vs SDataPro AIS 比對的 xlsx 轉成 SDataPro Report 系列 HTML 使用的 JSON 格式。

支援兩種 xlsx 格式：
  v1（舊版，含 SDataPro與船訊網兩者資料落差時間_分 欄位）：
    e.g. 20260515_船訊網_SDataPro資料比對.xlsx
  v2（新版，含 AI 補點 status 欄位）：
    e.g. 船訊網_SDataPro_AIS資料比較_20260519.xlsx

用法：
  python convert_xlsx_to_json.py <input.xlsx> <output.json> [--date YYYY-MM-DD]
  
例：
  python convert_xlsx_to_json.py data/船訊網_SDataPro_AIS資料比較_20260519.xlsx data/20260519.json --date 2026-05-19

需要套件：
  pip install openpyxl
"""

import argparse, json, math, os, sys, statistics
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("錯誤：缺少 openpyxl 套件。請執行 `pip install openpyxl`")
    sys.exit(1)


# =============================================================================
# 演算法常數（v3 / 工程師 generate_briefing.py 一致）
# =============================================================================
# 一致性閾值（XY vs SD 在這個誤差內視為「一致」）
ENG_POS_KM   = 5.0
ENG_HDG_DEG  = 10.0
ENG_COG_DEG  = 10.0
ENG_SOG_KN   = 2.0

# AIS 欄位有效範圍
ENG_NAVI_RANGE = (0, 15)
ENG_HDG_RANGE  = (0, 360)
ENG_COG_RANGE  = (0, 360)
ENG_SOG_RANGE  = (0, 102.3)
ENG_ROT_RANGE  = (-128, 127)

# 綜合分數加權（pos 主導，總和 1.0）
ENG_W_POS, ENG_W_HDG, ENG_W_COG, ENG_W_SOG, ENG_W_NAVI = 0.60, 0.10, 0.10, 0.10, 0.10

# 分級門檻（純分數帶）
ENG_PARTIAL_RATIO = 0.70    # sd_valid_ratio 低於此 → "部分資料"
ENG_QUAL_SCORE    = 90.0    # score >= → "A+B 可立即取代"
ENG_NEAR_SCORE    = 78.0    # score >= → "C 接近"，否則 "D 待改善"


def _haversine_km(lat1, lon1, lat2, lon2):
    """球面距離（公里）。輸入皆為 float。"""
    r = 6371.0088
    la1, lo1, la2, lo2 = map(math.radians, (lat1, lon1, lat2, lon2))
    dphi = la2 - la1
    dl = lo2 - lo1
    a = math.sin(dphi/2)**2 + math.cos(la1) * math.cos(la2) * math.sin(dl/2)**2
    return 2 * r * math.asin(min(1.0, math.sqrt(max(0.0, a))))


def _angular_diff(a, b):
    """環狀角度差（度，0~180）。"""
    x = abs(a - b) % 360
    return x if x <= 180 else 360 - x


def _in_range(x, lo, hi):
    return isinstance(x, (int, float)) and (lo <= x <= hi)


def _valid_lat(x):
    return isinstance(x, (int, float)) and -90 <= x <= 90 and x != 0


def _valid_lon(x):
    return isinstance(x, (int, float)) and -180 <= x <= 180 and x != 0


def detect_format(wb):
    """偵測 xlsx 格式 v1 / v2 / v3
    v3：只有 '資料源' + '狀態' 兩個工作表，所有統計都從 source 自算
    """
    if '狀態' in wb.sheetnames and '資料缺失統計' in wb.sheetnames:
        return 'v2'
    if 'source' in wb.sheetnames and 'mmsi與船名一覽' in wb.sheetnames:
        return 'v1'
    if 'source' in wb.sheetnames and '船名列表' in wb.sheetnames:
        return 'v2'
    if '資料源' in wb.sheetnames and '狀態' in wb.sheetnames:
        return 'v3'
    return 'unknown'


# 46 艘 YM 船 MMSI → (英文船名, 中文船名) 對照（用於 v3 等沒有船名表的格式）
FLEET_MAPPING = {
    416426000: ('YM HEIGHTS', '宇明輪'),
    416427000: ('YM HARMONY', '鎮明輪'),
    416428000: ('YM HORIZON', '環明輪'),
    416429000: ('YM HAWK', '威明輪'),
    416464000: ('YM UNICORN', '營明輪'),
    416465000: ('YM UPSURGENCE', '運明輪'),
    416466000: ('YM UNANIMITY', '續明輪'),
    416467000: ('YM UBIQUITY', '恆明輪'),
    416468000: ('YM UNIFORMITY', '結明輪'),
    416486000: ('YM INTELLIGENT', '精明輪'),
    416487000: ('YM INAUGURATION', '遠明輪'),
    416488000: ('YM IDEALS', '英明輪'),
    416490000: ('YM EVOLUTION', '創明輪'),
    416491000: ('YM ESSENCE', '實明輪'),
    416492000: ('YM IMMENSE', '雲明輪'),
    563155200: ('YM TROPHY', '納明輪'),
    563234400: ('YM CONTINUITY', '存明輪'),
    563234700: ('YM CAPACITY', '川明輪'),
    636012795: ('YM INCEPTION', '駿明輪'),
    636012796: ('YM IMAGE', '業明輪'),
    636012797: ('YM INITIATIVE', '暢明輪'),
    636013118: ('YM INSTRUCTION', '近明輪'),
    636013119: ('YM INTERACTION', '悅明輪'),
    636013121: ('YM IMPROVEMENT', '來明輪'),
    636013691: ('YM UPWARD', '常明輪'),
    636013692: ('YM UTILITY', '安明輪'),
    636013693: ('YM UNIFORM', '團明輪'),
    636013698: ('YM EFFICIENCY', '發明輪'),
    636013699: ('YM ETERNITY', '展明輪'),
    636014996: ('YM MUTUALITY', '震明輪'),
    636014997: ('YM MOBILITY', '動明輪'),
    636015182: ('YM MILESTONE', '盛明輪'),
    636015183: ('YM MASCULINITY', '鮮明輪'),
    636016703: ('YM ENLIGHTENMENT', '維明輪'),
    636016704: ('YM EXCELLENCE', '卓明輪'),
    636016705: ('YM EXPRESS', '越明輪'),
    636019893: ('YM CONTINENT', '洋明輪'),
    636019894: ('YM CREDENTIAL', '聚明輪'),
    636019895: ('YM CENTENNIAL', '百明輪'),
    636019897: ('YM CERTAINTY', '好明輪'),
    636019898: ('YM CREDIBILITY', '譽明輪'),
    636019899: ('YM CONSTANCY', '永明輪'),
    636019900: ('YM COOPERATION', '長明輪'),
    636021244: ('YM TOGETHER', '賢明輪'),
    636023581: ('YM WONDERLAND', '景明輪'),
    636023582: ('YM WISDOM', '承明輪'),
}


def downsample(pts, target=300):
    if len(pts) <= target:
        return pts[:]
    step = len(pts) / target
    return [pts[int(i * step)] for i in range(target)]


def median_minutes(times):
    if len(times) < 2:
        return None
    s = sorted(times)
    diffs = [(s[i+1] - s[i]).total_seconds() / 60 for i in range(len(s) - 1)]
    return statistics.median(diffs)


def process_v2(wb):
    """處理 v2 格式（含 AI 補點 status）"""
    # 船名列表
    ship_meta = {}
    for i, r in enumerate(wb['船名列表'].iter_rows(values_only=True)):
        if i == 0: continue
        ship_meta[r[0]] = {'en': r[1].strip() if r[1] else '', 'zh': r[2].strip() if r[2] else ''}
    
    # 資料缺失統計
    miss_stats = {}
    for i, r in enumerate(wb['資料缺失統計'].iter_rows(values_only=True)):
        if i == 0: continue
        miss_stats[r[0]] = {
            'n_total': r[1],
            'xy_missing': r[2],
            'sd_filled': r[3],
            'sd_failed': r[4],
            'sd_normal': r[5],
        }
    
    # 不合法比例統計
    invalid_stats = {}
    for i, r in enumerate(wb['各船AIS值不合法比例統計'].iter_rows(values_only=True)):
        if i == 0: continue
        invalid_stats[r[0]] = {
            'xy_inv': {
                'navi': r[5]*100, 'hdg': r[6]*100, 'cog': r[7]*100, 'sog': r[8]*100, 'rot': r[9]*100,
                'any': max(r[5], r[6], r[7], r[8], r[9]) * 100
            },
            'sd_inv': {
                'navi': r[12]*100, 'hdg': r[16]*100, 'cog': r[13]*100, 'sog': r[14]*100, 'rot': r[15]*100,
                'any': max(r[12], r[13], r[14], r[15], r[16]) * 100
            },
        }
    
    # P95 統計
    p95_stats = {}
    for i, r in enumerate(wb['所有船舶AIS資料差異P95'].iter_rows(values_only=True)):
        if i == 0: continue
        hdg_p95 = r[3] or 0
        cog_p95 = r[4] or 0
        sog_p95 = r[5] or 0
        lat_p95 = r[7] or 0
        lon_p95 = r[8] or 0
        pos_p95_km = ((lat_p95 * 111) ** 2 + (lon_p95 * 96) ** 2) ** 0.5
        p95_stats[r[0]] = {
            'hdg_p95': hdg_p95, 'cog_p95': cog_p95, 'sog_p95': sog_p95,
            'pos_p95_km': pos_p95_km, 'n_data': r[10] or 0,
        }
    
    # 一致率 + 軌跡 + 延遲（from source + AIS資料差異）
    consistency = defaultdict(lambda: {'pos_ok':0,'hdg_ok':0,'cog_ok':0,'sog_ok':0,'navi_ok':0,'total':0})
    for i, r in enumerate(wb['AIS資料差異'].iter_rows(values_only=True)):
        if i == 0: continue
        mmsi = r[0]
        hdg_diff = abs(r[1]) if r[1] is not None else 999
        cog_diff = abs(r[2]) if r[2] is not None else 999
        sog_diff = abs(r[3]) if r[3] is not None else 999
        lat_diff = abs(r[5]) if r[5] is not None else 999
        lon_diff = abs(r[6]) if r[6] is not None else 999
        navi_same = r[7]
        pos_diff_km = ((lat_diff*111)**2 + (lon_diff*96)**2)**0.5
        c = consistency[mmsi]
        c['total'] += 1
        if pos_diff_km <= 5.5: c['pos_ok'] += 1
        if hdg_diff <= 5: c['hdg_ok'] += 1
        if cog_diff <= 10: c['cog_ok'] += 1
        if sog_diff <= 1: c['sog_ok'] += 1
        if navi_same: c['navi_ok'] += 1
    
    # 軌跡 + 延遲（from source）
    # 注意：source 工作表的列順序是按 MMSI 而非時間，所以收集成 (time, point) tuple，
    # 後面再按時間排序，否則 polyline 會把點亂連成蜘蛛網。
    tracks_split = defaultdict(lambda: {
        'xy_pairs': [],          # list of (xy_t or None, [lat, lon])
        'sd_pairs': [],          # list of (sd_t or None, [lat, lon], status)
        'latencies': [],
        'createtimes': [],
    })
    for i, r in enumerate(wb['source'].iter_rows(values_only=True)):
        if i == 0: continue
        mmsi = r[0]
        xy_lat, xy_lon = r[2], r[3]
        xy_t = r[9]
        sd_lat, sd_lon = r[11], r[12]
        sd_t = r[18]
        status = r[20]
        createtime = r[21] if len(r) > 21 else None
        if xy_lat and xy_lon:
            tracks_split[mmsi]['xy_pairs'].append(
                (xy_t if isinstance(xy_t, datetime) else None, [xy_lat, xy_lon])
            )
        if sd_lat and sd_lon and status in (1, 2):
            tracks_split[mmsi]['sd_pairs'].append(
                (sd_t if isinstance(sd_t, datetime) else None, [sd_lat, sd_lon], status)
            )
        if isinstance(xy_t, datetime) and isinstance(sd_t, datetime):
            tracks_split[mmsi]['latencies'].append(abs((xy_t - sd_t).total_seconds())/60)
        if isinstance(createtime, datetime):
            tracks_split[mmsi]['createtimes'].append(createtime)

    # 對每艘船的點按時間排序，然後展開成原本下游使用的 xy/sd_normal/sd_filled/xy_times/sd_times
    _MAX_DT = datetime.max
    for mmsi, t in tracks_split.items():
        xy_sorted = sorted(t['xy_pairs'], key=lambda p: p[0] if p[0] is not None else _MAX_DT)
        sd_sorted = sorted(t['sd_pairs'], key=lambda p: p[0] if p[0] is not None else _MAX_DT)
        t['xy'] = [p[1] for p in xy_sorted]
        t['xy_times'] = [p[0] for p in xy_sorted if p[0] is not None]
        t['sd_normal'] = [p[1] for p in sd_sorted if p[2] == 1]
        t['sd_filled'] = [p[1] for p in sd_sorted if p[2] == 2]
        t['sd_times'] = [p[0] for p in sd_sorted if p[0] is not None]

    return ship_meta, miss_stats, invalid_stats, p95_stats, consistency, tracks_split


def process_v1(wb):
    """處理 v1 格式（舊版，無 AI 補點）"""
    # mmsi與船名一覽
    ship_meta = {}
    for i, r in enumerate(wb['mmsi與船名一覽'].iter_rows(values_only=True)):
        if i == 0: continue
        ship_meta[r[0]] = {'en': r[1].strip() if r[1] else '', 'zh': r[2].strip() if r[2] else ''}
    
    # source 表計算所有統計
    miss_stats = defaultdict(lambda: {'n_total':0,'xy_missing':0,'sd_filled':0,'sd_failed':0,'sd_normal':0})
    inv_count = defaultdict(lambda: {'xy':{'navi':0,'hdg':0,'cog':0,'sog':0,'rot':0,'any':0,'n':0},
                                      'sd':{'navi':0,'hdg':0,'cog':0,'sog':0,'rot':0,'any':0,'n':0}})
    consistency = defaultdict(lambda: {'pos_ok':0,'hdg_ok':0,'cog_ok':0,'sog_ok':0,'navi_ok':0,'total':0})
    p95_collect = defaultdict(lambda: {'pos':[],'hdg':[],'cog':[],'sog':[]})
    # 用 *_pairs 收集 (time, point) tuple，最後排序後再展開（修正 polyline 亂連的問題）
    tracks_split = defaultdict(lambda: {
        'xy_pairs': [], 'sd_pairs': [],
        'xy':[],'sd_normal':[],'sd_filled':[],'latencies':[],'xy_times':[],'sd_times':[]
    })
    
    for i, r in enumerate(wb['source'].iter_rows(values_only=True)):
        if i == 0: continue
        mmsi = r[0]
        ms = miss_stats[mmsi]
        ms['n_total'] += 1
        
        xy_lat, xy_lon, xy_navi, xy_hdg, xy_cog, xy_sog, xy_rot = r[2], r[3], r[4], r[5], r[6], r[7], r[8]
        sd_lat, sd_lon, sd_navi = r[12], r[13], r[14]
        sd_cog, sd_sog, sd_rot, sd_hdg = r[15], r[16], r[17], r[18]
        xy_t = r[9]; sd_t = r[19]
        
        # SD has data?
        sd_zero = (sd_lat is None or sd_lat == 0) and (sd_lon is None or sd_lon == 0)
        if sd_zero:
            ms['sd_failed'] += 1
        else:
            ms['sd_normal'] += 1
        
        # XY 不合法計數
        ix = inv_count[mmsi]['xy']
        ix['n'] += 1
        if xy_navi is None or not (0 <= xy_navi <= 15): ix['navi'] += 1
        if xy_hdg is None or xy_hdg > 360: ix['hdg'] += 1
        if xy_cog is None or xy_cog > 360: ix['cog'] += 1
        if xy_sog is None or xy_sog > 102.3: ix['sog'] += 1
        if xy_rot is None or not (-128 <= xy_rot <= 127): ix['rot'] += 1
        
        # SD 不合法計數（如果有資料；要小心 SD 欄位可能是字串如 'NULL'）
        if not sd_zero:
            ix2 = inv_count[mmsi]['sd']
            ix2['n'] += 1
            def is_num(x): return isinstance(x, (int, float))
            if not is_num(sd_navi) or not (0 <= sd_navi <= 15): ix2['navi'] += 1
            if not is_num(sd_hdg) or sd_hdg > 360: ix2['hdg'] += 1
            if not is_num(sd_cog) or sd_cog > 360: ix2['cog'] += 1
            if not is_num(sd_sog) or sd_sog > 102.3: ix2['sog'] += 1
            if not is_num(sd_rot) or not (-128 <= sd_rot <= 127): ix2['rot'] += 1
        
        # 一致率（僅當 SD 有資料）
        if not sd_zero and xy_lat and xy_lon:
            cs = consistency[mmsi]
            cs['total'] += 1
            lat_diff = abs(xy_lat - sd_lat)
            lon_diff = abs(xy_lon - sd_lon)
            pos_diff_km = ((lat_diff*111)**2 + (lon_diff*96)**2)**0.5
            if pos_diff_km <= 5.5: cs['pos_ok'] += 1
            def safe_diff(a, b, t):
                return isinstance(a, (int, float)) and isinstance(b, (int, float)) and abs(a - b) <= t
            if safe_diff(xy_hdg, sd_hdg, 5): cs['hdg_ok'] += 1
            if safe_diff(xy_cog, sd_cog, 10): cs['cog_ok'] += 1
            if safe_diff(xy_sog, sd_sog, 1): cs['sog_ok'] += 1
            if xy_navi == sd_navi: cs['navi_ok'] += 1
            
            p95_collect[mmsi]['pos'].append(pos_diff_km)
            if isinstance(xy_hdg, (int, float)) and isinstance(sd_hdg, (int, float)): p95_collect[mmsi]['hdg'].append(abs(xy_hdg - sd_hdg))
            if isinstance(xy_cog, (int, float)) and isinstance(sd_cog, (int, float)): p95_collect[mmsi]['cog'].append(abs(xy_cog - sd_cog))
            if isinstance(xy_sog, (int, float)) and isinstance(sd_sog, (int, float)): p95_collect[mmsi]['sog'].append(abs(xy_sog - sd_sog))

        
        # 軌跡（先存成 (time, point) tuple，迴圈外再排序）
        if xy_lat and xy_lon:
            tracks_split[mmsi]['xy_pairs'].append(
                (xy_t if isinstance(xy_t, datetime) else None, [xy_lat, xy_lon])
            )
        if not sd_zero:
            tracks_split[mmsi]['sd_pairs'].append(
                (sd_t if isinstance(sd_t, datetime) else None, [sd_lat, sd_lon], 1)
            )
        if isinstance(xy_t, datetime) and isinstance(sd_t, datetime) and not sd_zero:
            tracks_split[mmsi]['latencies'].append(abs((xy_t - sd_t).total_seconds())/60)
    
    # P95 calc
    p95_stats = {}
    for mmsi, pc in p95_collect.items():
        def p95(arr):
            if not arr: return 0
            arr.sort()
            return arr[int(len(arr)*0.95)]
        p95_stats[mmsi] = {
            'pos_p95_km': p95(pc['pos']),
            'hdg_p95': p95(pc['hdg']),
            'cog_p95': p95(pc['cog']),
            'sog_p95': p95(pc['sog']),
            'n_data': len(pc['pos']),
        }
    
    # invalid as percentages
    invalid_stats = {}
    for mmsi, c in inv_count.items():
        def pct_dict(d):
            n = d['n'] or 1
            r = {k: d[k]/n*100 for k in ('navi','hdg','cog','sog','rot')}
            r['any'] = max(r.values())
            return r
        invalid_stats[mmsi] = {'xy_inv': pct_dict(c['xy']), 'sd_inv': pct_dict(c['sd']) if c['sd']['n'] > 0 else None}

    # v1 軌跡也按時間排序（修正 polyline 亂連）
    _MAX_DT_V1 = datetime.max
    for mmsi, t in tracks_split.items():
        xy_sorted = sorted(t['xy_pairs'], key=lambda p: p[0] if p[0] is not None else _MAX_DT_V1)
        sd_sorted = sorted(t['sd_pairs'], key=lambda p: p[0] if p[0] is not None else _MAX_DT_V1)
        t['xy'] = [p[1] for p in xy_sorted]
        t['xy_times'] = [p[0] for p in xy_sorted if p[0] is not None]
        t['sd_normal'] = [p[1] for p in sd_sorted]  # v1 only has status 1 (no AI fill)
        t['sd_times'] = [p[0] for p in sd_sorted if p[0] is not None]

    return ship_meta, dict(miss_stats), invalid_stats, p95_stats, dict(consistency), dict(tracks_split)


def process_v3(wb):
    """處理 v3 格式：只有 '資料源' + '狀態' 兩個工作表。
    演算法與工程師 generate_briefing.py 一致：
      - 一致率以 haversine + angular_diff 計算
      - 各欄位的分母為「雙邊皆有效」的列數（per-field denominator）
      - 不合法率以 AIS 有效範圍判定
      - sd_valid = 一列裡 7 個 SD 欄位 + 經緯度 全在有效範圍
      - 軌跡按 createtime 排序

    status: 1=都正常, 2=sd 異常但 slab 正常(AI 補點), 3=xy 異常 sd 正常,
            4=xy/sd 都異常但 slab 正常, 5=全部都壞
    """
    ship_meta = {mmsi: {'en': en, 'zh': zh} for mmsi, (en, zh) in FLEET_MAPPING.items()}

    miss_stats = defaultdict(lambda: {'n_total':0,'xy_missing':0,'sd_filled':0,'sd_failed':0,'sd_normal':0,'sd_valid':0})
    inv_count = defaultdict(lambda: {'xy':{'navi':0,'hdg':0,'cog':0,'sog':0,'rot':0,'any':0,'n':0},
                                      'sd':{'navi':0,'hdg':0,'cog':0,'sog':0,'rot':0,'any':0,'n':0}})
    # 工程師版：每個欄位獨立分母（雙邊皆有效時才計入）
    consistency = defaultdict(lambda: {
        'pos_ok':0, 'pos_total':0,
        'pos_ok_adj':0,       # 延遲補償後的 pos_ok（扣除船訊網延遲導致的位置漂移）
        'hdg_ok':0, 'hdg_total':0,
        'cog_ok':0, 'cog_total':0,
        'sog_ok':0, 'sog_total':0,
        'navi_ok':0, 'navi_total':0,
        'total':0,            # n_total（含全部列）
    })
    p95_collect = defaultdict(lambda: {'pos':[],'hdg':[],'cog':[],'sog':[]})
    tracks_split = defaultdict(lambda: {
        'xy_pairs': [], 'sd_pairs': [],
        'xy':[],'sd_normal':[],'sd_filled':[],'latencies':[],'xy_times':[],'sd_times':[],
        'createtimes': [],
    })

    for i, r in enumerate(wb['資料源'].iter_rows(values_only=True)):
        if i == 0: continue
        mmsi = r[0]
        if mmsi is None: continue
        ms = miss_stats[mmsi]
        ms['n_total'] += 1
        cs = consistency[mmsi]
        cs['total'] += 1

        # v2/v3 欄位佈局
        xy_lat, xy_lon, xy_navi = r[2], r[3], r[4]
        xy_hdg, xy_cog, xy_sog, xy_rot = r[5], r[6], r[7], r[8]
        xy_t = r[9]
        sd_lat, sd_lon, sd_navi = r[11], r[12], r[13]
        sd_cog, sd_sog, sd_rot, sd_hdg = r[14], r[15], r[16], r[17]
        sd_t = r[18]
        status = r[20]
        createtime = r[21] if len(r) > 21 else None

        # 依 status 歸類 SD 資料來源（保留 AI 補點追蹤，用於 Full Report 顯示）
        if status == 1:
            ms['sd_normal'] += 1
        elif status == 2:
            ms['sd_filled'] += 1
        elif status == 3:
            ms['sd_normal'] += 1
        elif status == 4:
            ms['sd_filled'] += 1
        else:
            ms['sd_failed'] += 1

        # 各欄位有效性
        xy_lat_ok = _valid_lat(xy_lat); xy_lon_ok = _valid_lon(xy_lon)
        sd_lat_ok = _valid_lat(sd_lat); sd_lon_ok = _valid_lon(sd_lon)
        xy_pos_ok = xy_lat_ok and xy_lon_ok
        sd_pos_ok = sd_lat_ok and sd_lon_ok
        xy_navi_ok = _in_range(xy_navi, *ENG_NAVI_RANGE); sd_navi_ok = _in_range(sd_navi, *ENG_NAVI_RANGE)
        xy_hdg_ok  = _in_range(xy_hdg,  *ENG_HDG_RANGE);  sd_hdg_ok  = _in_range(sd_hdg,  *ENG_HDG_RANGE)
        xy_cog_ok  = _in_range(xy_cog,  *ENG_COG_RANGE);  sd_cog_ok  = _in_range(sd_cog,  *ENG_COG_RANGE)
        xy_sog_ok  = _in_range(xy_sog,  *ENG_SOG_RANGE);  sd_sog_ok  = _in_range(sd_sog,  *ENG_SOG_RANGE)
        xy_rot_ok  = _in_range(xy_rot,  *ENG_ROT_RANGE);  sd_rot_ok  = _in_range(sd_rot,  *ENG_ROT_RANGE)

        # SD 整列有效（用於 sd_valid_ratio → partial 判定）
        sd_all_valid = sd_pos_ok and sd_navi_ok and sd_hdg_ok and sd_cog_ok and sd_sog_ok and sd_rot_ok
        if sd_all_valid:
            ms['sd_valid'] += 1

        # XY 不合法（依 AIS 有效範圍）
        ix = inv_count[mmsi]['xy']
        ix['n'] += 1
        if not xy_navi_ok: ix['navi'] += 1
        if not xy_hdg_ok:  ix['hdg']  += 1
        if not xy_cog_ok:  ix['cog']  += 1
        if not xy_sog_ok:  ix['sog']  += 1
        if not xy_rot_ok:  ix['rot']  += 1

        # SD 不合法（全列都算，不過濾零座標）
        ix2 = inv_count[mmsi]['sd']
        ix2['n'] += 1
        if not sd_navi_ok: ix2['navi'] += 1
        if not sd_hdg_ok:  ix2['hdg']  += 1
        if not sd_cog_ok:  ix2['cog']  += 1
        if not sd_sog_ok:  ix2['sog']  += 1
        if not sd_rot_ok:  ix2['rot']  += 1

        # 一致率（per-field 分母：雙邊皆有效時才計入）
        if xy_pos_ok and sd_pos_ok:
            cs['pos_total'] += 1
            dist_km = _haversine_km(float(xy_lat), float(xy_lon), float(sd_lat), float(sd_lon))
            if dist_km <= ENG_POS_KM: cs['pos_ok'] += 1
            p95_collect[mmsi]['pos'].append(dist_km)
            # 延遲補償：扣除「兩家系統時間差 × 船速」所造成的預期位置漂移
            # 預期漂移 = |xy_t − sd_t| (min) × sog (knot) × 1.852/60 (km/min/knot)
            # adjusted = max(0, actual − expected_drift) ≤ 5 km 即視為「真正一致」
            adj_dist = dist_km
            if isinstance(xy_t, datetime) and isinstance(sd_t, datetime):
                delay_min = abs((xy_t - sd_t).total_seconds()) / 60.0
                # 取 XY 與 SD 兩端 sog 平均當船速估計（兩家都應該記錄相同瞬時速度）
                speeds = [s for s in (xy_sog, sd_sog) if _in_range(s, *ENG_SOG_RANGE)]
                sog_kn = (sum(speeds) / len(speeds)) if speeds else 0.0
                expected_drift_km = delay_min * sog_kn * (1.852 / 60.0)
                adj_dist = max(0.0, dist_km - expected_drift_km)
            if adj_dist <= ENG_POS_KM: cs['pos_ok_adj'] += 1
        if xy_hdg_ok and sd_hdg_ok:
            cs['hdg_total'] += 1
            d = _angular_diff(float(xy_hdg), float(sd_hdg))
            if d <= ENG_HDG_DEG: cs['hdg_ok'] += 1
            p95_collect[mmsi]['hdg'].append(d)
        if xy_cog_ok and sd_cog_ok:
            cs['cog_total'] += 1
            d = _angular_diff(float(xy_cog), float(sd_cog))
            if d <= ENG_COG_DEG: cs['cog_ok'] += 1
            p95_collect[mmsi]['cog'].append(d)
        if xy_sog_ok and sd_sog_ok:
            cs['sog_total'] += 1
            d = abs(float(xy_sog) - float(sd_sog))
            if d <= ENG_SOG_KN: cs['sog_ok'] += 1
            p95_collect[mmsi]['sog'].append(d)
        if xy_navi_ok and sd_navi_ok:
            cs['navi_total'] += 1
            if xy_navi == sd_navi: cs['navi_ok'] += 1

        # 軌跡（status 1 → sd_normal, 2 → sd_filled；按 createtime 排序）
        if xy_lat and xy_lon:
            tracks_split[mmsi]['xy_pairs'].append(
                (createtime if isinstance(createtime, datetime) else (xy_t if isinstance(xy_t, datetime) else None),
                 [xy_lat, xy_lon])
            )
        if sd_lat and sd_lon and status in (1, 2):
            tracks_split[mmsi]['sd_pairs'].append(
                (createtime if isinstance(createtime, datetime) else (sd_t if isinstance(sd_t, datetime) else None),
                 [sd_lat, sd_lon], status)
            )
        if isinstance(xy_t, datetime) and isinstance(sd_t, datetime):
            tracks_split[mmsi]['latencies'].append(abs((xy_t - sd_t).total_seconds())/60)
        if isinstance(createtime, datetime):
            tracks_split[mmsi]['createtimes'].append(createtime)

    # 按時間排序軌跡（修正 polyline 蜘蛛網）
    _MAX_DT = datetime.max
    for mmsi, t in tracks_split.items():
        xy_sorted = sorted(t['xy_pairs'], key=lambda p: p[0] if p[0] is not None else _MAX_DT)
        sd_sorted = sorted(t['sd_pairs'], key=lambda p: p[0] if p[0] is not None else _MAX_DT)
        t['xy'] = [p[1] for p in xy_sorted]
        t['xy_times'] = [p[0] for p in xy_sorted if p[0] is not None]
        t['sd_normal'] = [p[1] for p in sd_sorted if p[2] == 1]
        t['sd_filled'] = [p[1] for p in sd_sorted if p[2] == 2]
        t['sd_times'] = [p[0] for p in sd_sorted if p[0] is not None]

    # P95 計算
    p95_stats = {}
    for mmsi, pc in p95_collect.items():
        def p95(arr):
            if not arr: return 0
            arr.sort()
            return arr[min(len(arr)-1, int(len(arr)*0.95))]
        p95_stats[mmsi] = {
            'pos_p95_km': p95(pc['pos']),
            'hdg_p95': p95(pc['hdg']),
            'cog_p95': p95(pc['cog']),
            'sog_p95': p95(pc['sog']),
            'n_data': len(pc['pos']),
        }

    # 不合法率轉百分比
    invalid_stats = {}
    for mmsi, c in inv_count.items():
        def pct_dict(d):
            n = d['n'] or 1
            r = {k: round(d[k]/n*100, 4) for k in ('navi','hdg','cog','sog','rot')}
            r['any'] = round(max(r.values()), 4)
            return r
        invalid_stats[mmsi] = {
            'xy_inv': pct_dict(c['xy']),
            'sd_inv': pct_dict(c['sd']) if c['sd']['n'] > 0 else None,
        }

    return ship_meta, dict(miss_stats), invalid_stats, p95_stats, dict(consistency), dict(tracks_split)


def build_vessels(ship_meta, miss_stats, invalid_stats, p95_stats, consistency, tracks_split):
    vessels = []
    latency_med = {}
    xy_int = {}
    sd_int = {}
    
    for m, t in tracks_split.items():
        if t['latencies']:
            latency_med[m] = round(statistics.median(t['latencies']))
        xy_int[m] = round(median_minutes(t['xy_times'])) if median_minutes(t['xy_times']) is not None else None
        sd_int[m] = round(median_minutes(t['sd_times'])) if median_minutes(t['sd_times']) is not None else None
    
    for mmsi in sorted(ship_meta.keys()):
        meta = ship_meta[mmsi]
        ms = miss_stats.get(mmsi, {})
        n_total = ms.get('n_total', 0)
        sd_normal = ms.get('sd_normal', 0)
        sd_filled = ms.get('sd_filled', 0)
        sd_failed = ms.get('sd_failed', 0)
        sd_valid_status = sd_normal + sd_filled               # 依 status 分類的可用筆數（含 AI 補點）
        sd_valid_eng = ms.get('sd_valid', sd_valid_status)    # 工程師演算法：所有 SD 欄位都在有效範圍的筆數

        inv = invalid_stats.get(mmsi, {'xy_inv': {'navi':0,'hdg':0,'cog':0,'sog':0,'rot':0,'any':0}, 'sd_inv': None})
        p95 = p95_stats.get(mmsi, {})
        cons = consistency.get(mmsi, {'total':0,'pos_ok':0,'hdg_ok':0,'cog_ok':0,'sog_ok':0,'navi_ok':0})

        # 偵測是否為 v3（per-field 分母）；否則退回舊邏輯（共用 total 分母）
        engineer_algo = 'pos_total' in cons

        if engineer_algo:
            # ── 工程師演算法 ──
            sd_valid_ratio = sd_valid_eng / n_total if n_total > 0 else 0
            is_no_data = sd_valid_ratio <= 0
            has_partial = (not is_no_data) and (sd_valid_ratio < ENG_PARTIAL_RATIO)

            def _rate(ok, denom):
                return round(100 * ok / denom, 2) if denom > 0 else 0
            comp = {
                'pos':      _rate(cons['pos_ok'],  cons.get('pos_total', 0)),
                'pos_adj':  _rate(cons.get('pos_ok_adj', 0), cons.get('pos_total', 0)),
                'hdg':      _rate(cons['hdg_ok'],  cons.get('hdg_total', 0)),
                'cog':      _rate(cons['cog_ok'],  cons.get('cog_total', 0)),
                'sog':      _rate(cons['sog_ok'],  cons.get('sog_total', 0)),
                'navistat': _rate(cons['navi_ok'], cons.get('navi_total', 0)),
            }
            p95_out = {
                'pos': round(p95.get('pos_p95_km', 0), 3),
                'hdg': p95.get('hdg_p95'),
                'cog': p95.get('cog_p95'),
                'sog': round(p95.get('sog_p95', 0), 3),
            }
            quality = (comp['pos']*ENG_W_POS + comp['hdg']*ENG_W_HDG + comp['cog']*ENG_W_COG
                       + comp['sog']*ENG_W_SOG + comp['navistat']*ENG_W_NAVI)
            score = round(quality * sd_valid_ratio, 2)
            n_real = sd_valid_status
        else:
            # ── 舊（v1/v2）演算法 ──
            n_real = sd_valid_status
            is_no_data = (sd_normal == 0 and sd_filled == 0)
            has_partial = (n_real > 0) and (n_real < n_total) and not is_no_data
            t = cons.get('total', 0) or 1
            if is_no_data:
                comp = {'pos':0,'hdg':0,'cog':0,'sog':0,'navistat':0}
                p95_out = {'pos':None,'hdg':None,'cog':None,'sog':None}
                score = 0
            else:
                comp = {
                    'pos': round(100*cons['pos_ok']/t, 2),
                    'hdg': round(100*cons['hdg_ok']/t, 2),
                    'cog': round(100*cons['cog_ok']/t, 2),
                    'sog': round(100*cons['sog_ok']/t, 2),
                    'navistat': round(100*cons['navi_ok']/t, 2),
                }
                p95_out = {
                    'pos': round(p95.get('pos_p95_km', 0), 3),
                    'hdg': p95.get('hdg_p95'),
                    'cog': p95.get('cog_p95'),
                    'sog': round(p95.get('sog_p95', 0), 3),
                }
                quality = comp['pos']*0.35 + comp['hdg']*0.20 + comp['cog']*0.10 + comp['sog']*0.20 + comp['navistat']*0.15
                coverage = n_real / n_total if n_total > 0 else 0
                score = round(quality * coverage, 2)

        vessels.append({
            'en': meta['en'], 'zh': meta['zh'],
            'n_total': n_total, 'n_zero': sd_failed, 'n_real': n_real,
            'is_no_data': is_no_data, 'has_partial': has_partial,
            'sd_note': 'no_data' if is_no_data else ('partial' if has_partial else 'full'),
            'comp': comp, 'p95': p95_out, 'score': score,
            'navistat_a1': comp['navistat'], 'update_count': n_real,
            'xy_interval_med': xy_int.get(mmsi),
            'sd_interval_med': sd_int.get(mmsi),
            'latency_median': latency_med.get(mmsi),
            'latency_p95': None,
            'xy_inv': inv['xy_inv'],
            'sd_inv': inv['sd_inv'] if not is_no_data else None,
            'sd_filled_count': sd_filled,
            'sd_normal_count': sd_normal,
            'sd_failed_count': sd_failed,
        })
    return vessels


def build_tracks(tracks_split, ship_meta):
    new_tracks = {}
    for mmsi, t in tracks_split.items():
        if mmsi not in ship_meta: continue
        en = ship_meta[mmsi]['en']
        new_tracks[en] = {
            'xy': downsample(t['xy']),
            'sd': downsample(t['sd_normal'] + t['sd_filled']),
            'sd_normal': downsample(t['sd_normal']),
            'sd_filled': downsample(t['sd_filled']),
        }
    return new_tracks


def categorize(vessels):
    """分級規則（與工程師 generate_briefing.py 一致）：
      - 無資料：is_no_data
      - 部分資料：has_partial（v3：sd_valid_ratio<0.70；v1/v2：n_real<n_total）
      - A+B 可立即取代：score >= ENG_QUAL_SCORE (90)
      - C 接近：score >= ENG_NEAR_SCORE (78)
      - D 待改善：其餘
    """
    qual, partial_s, near, unqual, no_data = [], [], [], [], []
    for v in vessels:
        if v['is_no_data']:
            no_data.append(v['en'])
        elif v['has_partial']:
            partial_s.append(v['en'])
        elif v['score'] >= ENG_QUAL_SCORE:
            qual.append(v['en'])
        elif v['score'] >= ENG_NEAR_SCORE:
            near.append(v['en'])
        else:
            unqual.append(v['en'])
    return {'qual': qual, 'near': near, 'partial': partial_s, 'unqual': unqual, 'nodata': no_data}



def find_main_period(times):
    """找出實際連續資料的主要區間（過濾零星散落的資料點）
    
    Returns (start_date, end_date, n_days, sparse_dates)
    其中 sparse_dates 是被排除的零星日期清單
    """
    from collections import Counter
    from datetime import timedelta
    if not times:
        return None, None, 0, []
    by_date = Counter(t.date() for t in times)
    sorted_dates = sorted(by_date.keys())
    total = sum(by_date.values())
    
    # 把「資料量 < 總量 5%」的日子視為零星
    threshold = max(total * 0.05, 100)
    substantial = [d for d in sorted_dates if by_date[d] >= threshold]
    sparse = [d for d in sorted_dates if by_date[d] < threshold]
    
    if not substantial:
        return sorted_dates[0], sorted_dates[-1], len(sorted_dates), []
    
    # 找出 substantial 中的最長連續區間
    best_run = [substantial[0]]
    current = [substantial[0]]
    for i in range(1, len(substantial)):
        if substantial[i] == substantial[i-1] + timedelta(days=1):
            current.append(substantial[i])
        else:
            if len(current) > len(best_run):
                best_run = current
            current = [substantial[i]]
    if len(current) > len(best_run):
        best_run = current
    
    return best_run[0], best_run[-1], len(best_run), sparse


def main():
    ap = argparse.ArgumentParser(description='Convert AIS comparison xlsx to JSON for SDataPro Reports')
    ap.add_argument('input', help='Input xlsx file path')
    ap.add_argument('output', help='Output JSON file path')
    ap.add_argument('--date', help='Report date YYYY-MM-DD (auto-detected if omitted)')
    args = ap.parse_args()
    
    if not os.path.exists(args.input):
        print(f"錯誤：找不到輸入檔案 {args.input}")
        sys.exit(1)
    
    print(f"讀取 {args.input}...")
    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    fmt = detect_format(wb)
    print(f"偵測到格式：{fmt}")
    if fmt == 'unknown':
        print("錯誤：不支援的 xlsx 格式")
        sys.exit(1)
    
    if fmt == 'v2':
        ship_meta, miss_stats, invalid_stats, p95_stats, consistency, tracks_split = process_v2(wb)
    elif fmt == 'v3':
        ship_meta, miss_stats, invalid_stats, p95_stats, consistency, tracks_split = process_v3(wb)
    else:
        ship_meta, miss_stats, invalid_stats, p95_stats, consistency, tracks_split = process_v1(wb)
    
    print(f"處理 {len(ship_meta)} 艘船...")
    vessels = build_vessels(ship_meta, miss_stats, invalid_stats, p95_stats, consistency, tracks_split)
    tracks = build_tracks(tracks_split, ship_meta)
    cats = categorize(vessels)
    
    # AI 船舶清單
    ai_ships = []
    for v in vessels:
        if v.get('sd_filled_count', 0) > 0:
            fill_pct = round(v['sd_filled_count'] / v['n_total'] * 100, 1) if v['n_total'] > 0 else 0
            ai_ships.append({'en': v['en'], 'zh': v['zh'], 'fill_pct': fill_pct, 'count': v['sd_filled_count']})
    ai_ships.sort(key=lambda x: -x['fill_pct'])
    
    # 時間範圍（優先使用 createtime = 真正的資料收集時間；無 createtime 則退回 xy_lasttime_fd）
    all_xy = []
    for t in tracks_split.values():
        ct = t.get('createtimes', [])
        if ct:
            all_xy.extend(ct)
        else:
            all_xy.extend(t.get('xy_times', []))
    sparse_info = ""
    if all_xy:
        main_start, main_end, main_days, sparse = find_main_period(all_xy)
        # 主要連續區間用於顯示
        period_start = main_start.strftime('%Y/%m/%d').replace('/0', '/')
        period_end = main_end.strftime('%Y/%m/%d').replace('/0', '/')
        period_days = main_days
        # 找該主要區間內的真實時戳邊界
        in_main = [t for t in all_xy if main_start <= t.date() <= main_end]
        if in_main:
            _s, _e = min(in_main), max(in_main)
            period_label = f"{_s.year}/{_s.month}/{_s.day} {_s:%H:%M}-{_e.year}/{_e.month}/{_e.day} {_e:%H:%M}"
        else:
            period_label = f"{period_start}-{period_end}"
        # 散落資料註記
        if sparse:
            sparse_strs = [f"{d.month}/{d.day}" for d in sparse]
            sparse_info = f"另有 {sum(1 for t in all_xy if t.date() in sparse)} 筆零星資料分散於 {', '.join(sparse_strs)}（不計入主要區間）"
        # 總跨距資訊（含散落）
        total_span_start = min(all_xy).strftime('%Y/%m/%d').replace('/0', '/')
        total_span_end = max(all_xy).strftime('%Y/%m/%d').replace('/0', '/')
    else:
        period_start = period_end = '?'
        period_label = '?'
        period_days = 0
        total_span_start = total_span_end = '?'
    
    # Global aggregates
    total_pos_ok = sum(consistency[m]['pos_ok'] for m in consistency)
    total_hdg_ok = sum(consistency[m]['hdg_ok'] for m in consistency)
    total_cog_ok = sum(consistency[m]['cog_ok'] for m in consistency)
    total_sog_ok = sum(consistency[m]['sog_ok'] for m in consistency)
    total_navi_ok = sum(consistency[m]['navi_ok'] for m in consistency)
    total = sum(consistency[m]['total'] for m in consistency) or 1
    
    n_normal = sum(v.get('sd_normal_count', 0) for v in vessels)
    n_filled = sum(v.get('sd_filled_count', 0) for v in vessels)
    n_valid = sum(v['n_real'] for v in vessels if not v['is_no_data'])
    n_total_records = sum(v['n_total'] for v in vessels)
    
    out = {
        'meta': {
            'date_label': args.date or period_end.replace('/', '-'),
            'period_label': period_label,
            'period_start': period_start,
            'period_end': period_end,
            'period_days': period_days,
            'sparse_info': sparse_info,
            'total_span_start': total_span_start,
            'total_span_end': total_span_end,
            'report_date': args.date or period_end,
            'data_source': os.path.basename(args.input),
            'has_ai_fill': n_filled > 0,
            'n_ships': len(vessels),
            'n_total_records': n_total_records,
            'n_valid_records': n_valid,
            'n_sd_normal': n_normal,
            'n_ai_filled': n_filled,
        },
        'global': {
            'n_vessels_included': len(vessels) - sum(1 for v in vessels if v['is_no_data']),
            'n_vessels_excluded': sum(1 for v in vessels if v['is_no_data']),
            'n_rows': n_valid,
            'pos': round(100*total_pos_ok/total, 2),
            'hdg': round(100*total_hdg_ok/total, 2),
            'cog': round(100*total_cog_ok/total, 2),
            'sog': round(100*total_sog_ok/total, 2),
            'navistat': round(100*total_navi_ok/total, 2),
        },
        'vessels': vessels,
        'tracks': tracks,
        'categories': cats,
        'ai_ships': ai_ships,
    }
    
    print(f"寫入 {args.output}...")
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
    
    print("[OK] 完成！")
    print(f"  - 船舶數: {len(vessels)}")
    print(f"  - 期間: {period_start} ~ {period_end} ({period_days} 天)")
    print(f"  - 有效記錄: {n_valid:,}")
    print(f"  - AI 補點: {n_filled:,}" if n_filled > 0 else "  - 無 AI 補點資料")
    print(f"  - 分類: A+B {len(cats['qual'])} | C {len(cats['near'])} | 部分 {len(cats['partial'])} | D {len(cats['unqual'])} | 無資料 {len(cats['nodata'])}")


if __name__ == '__main__':
    main()
